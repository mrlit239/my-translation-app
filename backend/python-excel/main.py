"""
Excel Translation Service with Streaming Progress
Translates Excel files while preserving images, charts, and formatting using openpyxl
"""

import os
import io
import re
import json
import asyncio
from typing import Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import google.generativeai as genai
from dotenv import load_dotenv

load_dotenv()

app = FastAPI(title="Excel Translation Service")

# CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure Gemini
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)


def has_translatable_text(text: str) -> bool:
    """Check if text contains Japanese or Chinese characters"""
    if not text or not isinstance(text, str):
        return False
    pattern = r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FAF\u3400-\u4DBF]'
    return bool(re.search(pattern, text))


def extract_translatable_cells(wb) -> list:
    """Extract all cells with Japanese/Chinese text from workbook"""
    cells = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        if not isinstance(sheet, Worksheet):
            print(f"Skipping non-worksheet: {sheet_name}")
            continue
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and has_translatable_text(cell.value):
                    cells.append({
                        'sheet': sheet_name,
                        'coord': cell.coordinate,
                        'value': cell.value
                    })
    
    return cells


def translate_batch_sync(texts: list, prompt: str, model_name: str, api_key: str) -> dict:
    """Translate a batch of texts using Gemini (synchronous)"""
    key_to_use = api_key or GEMINI_API_KEY
    if not key_to_use:
        return {}
    
    genai.configure(api_key=key_to_use)
    
    batch_text = "\n".join([f"【{i}】{text}" for i, text in enumerate(texts)])
    
    full_prompt = f"""{prompt}

EXCEL CELL TRANSLATION RULES:
- Each cell is marked with 【number】
- Translate each cell and KEEP the 【number】 markers
- Keep translations concise (spreadsheet cells)
- Preserve numbers, dates, file names
- Output: 【0】translation【1】translation...

Text:
{batch_text}"""
    
    try:
        model = genai.GenerativeModel(model_name)
        response = model.generate_content(full_prompt)
        result_text = response.text
        
        pattern = r'【(\d+)】([^【]*)'
        matches = re.findall(pattern, result_text)
        
        translations = {}
        for idx, text in matches:
            translations[int(idx)] = text.strip()
        
        return translations
    except Exception as e:
        print(f"[ERROR] Translation error: {e}")
        return {}


@app.get("/health")
async def health():
    return {"status": "ok", "service": "excel-translation"}


@app.post("/translate-excel-stream")
async def translate_excel_stream(
    file: UploadFile = File(...),
    prompt: str = Form(default="Translate Japanese/Chinese to Vietnamese. Professional tone."),
    model: str = Form(default="gemini-2.0-flash"),
    api_key: str = Form(default="")
):
    """
    Translate Excel with Server-Sent Events progress updates.
    Returns NDJSON stream with progress, then final file as base64.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file")
    
    content = await file.read()
    
    async def generate():
        try:
            # Load workbook
            wb = load_workbook(io.BytesIO(content))
            cells = extract_translatable_cells(wb)
            total_cells = len(cells)
            
            yield json.dumps({"type": "start", "total": total_cells, "sheets": wb.sheetnames}) + "\n"
            
            if not cells:
                yield json.dumps({"type": "info", "message": "No translatable cells found"}) + "\n"
            else:
                # Batch translate with progress
                batch_size = 10
                translated_count = 0
                
                for i in range(0, len(cells), batch_size):
                    batch = cells[i:i + batch_size]
                    texts = [c['value'] for c in batch]
                    
                    # Translate this batch
                    translations = translate_batch_sync(texts, prompt, model, api_key)
                    
                    # Apply translations
                    for j, cell_info in enumerate(batch):
                        if j in translations:
                            sheet = wb[cell_info['sheet']]
                            sheet[cell_info['coord']] = translations[j]
                    
                    translated_count += len(batch)
                    percent = int((translated_count / total_cells) * 100)
                    
                    yield json.dumps({
                        "type": "progress",
                        "current": translated_count,
                        "total": total_cells,
                        "percent": percent,
                        "batch": i // batch_size + 1,
                        "sample": translations.get(0, "")[:30] if translations else ""
                    }) + "\n"
                    
                    # Small delay to prevent overwhelming
                    await asyncio.sleep(0.1)
            
            # Save workbook
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Encode as base64
            import base64
            file_data = base64.b64encode(output.read()).decode('utf-8')
            
            from urllib.parse import quote
            filename = file.filename.replace('.xlsx', '_translated.xlsx')
            
            yield json.dumps({
                "type": "complete",
                "filename": filename,
                "data": file_data
            }) + "\n"
            
        except Exception as e:
            yield json.dumps({"type": "error", "message": str(e)}) + "\n"
    
    return StreamingResponse(
        generate(),
        media_type="application/x-ndjson",
        headers={"X-Content-Type-Options": "nosniff"}
    )


# Keep the old endpoint for backwards compatibility
@app.post("/translate-excel")
async def translate_excel(
    file: UploadFile = File(...),
    prompt: str = Form(default="Translate Japanese/Chinese to Vietnamese. Professional tone."),
    model: str = Form(default="gemini-2.0-flash"),
    api_key: str = Form(default="")
):
    """Non-streaming version for compatibility"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file")
    
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        cells = extract_translatable_cells(wb)
        
        print(f"[DEBUG] Found {len(cells)} cells to translate")
        
        if cells:
            batch_size = 10
            for i in range(0, len(cells), batch_size):
                batch = cells[i:i + batch_size]
                texts = [c['value'] for c in batch]
                
                print(f"[DEBUG] Translating batch {i//batch_size + 1}/{(len(cells)-1)//batch_size + 1}")
                
                translations = translate_batch_sync(texts, prompt, model, api_key)
                
                for j, cell_info in enumerate(batch):
                    if j in translations:
                        sheet = wb[cell_info['sheet']]
                        sheet[cell_info['coord']] = translations[j]
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from urllib.parse import quote
        filename = quote(file.filename.replace('.xlsx', '_translated.xlsx'))
        
        print(f"[DEBUG] Translation complete! Sending file.")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
        
    except Exception as e:
        print(f"[ERROR] {e}")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
